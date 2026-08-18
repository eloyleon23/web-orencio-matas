#!/usr/bin/env python3
"""
Genera sugerencias de productos relacionados ("compra conjunta"), para
revisar y validar antes de pegarlas en la columna "relacionados" de la
hoja Productos del Sheet.

NO escribe nada directamente en el Sheet ni en productos.json — genera
un Excel de revisión. Tú decides qué sugerencias aceptar, editando el
Excel o directamente la columna "relacionados" del Sheet.

Uso:
    python3 generar_sugerencias_relacionados.py --entrada productos.json --salida sugerencias_relacionados.xlsx

El archivo productos.json de entrada es el mismo que ya genera el
proyecto (descárgalo de data/productos.json en el repositorio, o de la
caché de Drive vía el Web App de Apps Script).

Las reglas están pensadas para droguería, perfumería y pinturas — son
las áreas donde el emparejamiento por familia/subfamilia funciona bien.
Talleres queda fuera por ahora (apenas tiene subfamilias informadas en
el Sheet; necesitaría un enfoque distinto, más basado en el nombre).

Cada regla es deliberadamente conservadora: solo sugiere cuando
encuentra una coincidencia razonablemente clara (por subfamilia
exacta o por palabra clave inequívoca en el nombre) — se prefiere NO
sugerir nada antes que sugerir algo que no tenga sentido. Aun así,
TODO lo que genera esta herramienta es una sugerencia a revisar, no
una verdad asumida.
"""
import argparse
import json
import re
import unicodedata
from collections import defaultdict

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError:
    raise SystemExit("Falta openpyxl. Instala con: pip install openpyxl")


def sin_acentos_mayus(texto):
    """Mismo criterio que quitarAcentosMay() en buscador.html — mayúsculas
    sin acentos, para comparar de forma consistente con el resto del
    proyecto."""
    if not texto:
        return ''
    nfkd = unicodedata.normalize('NFKD', texto)
    return ''.join(c for c in nfkd if not unicodedata.combining(c)).upper()


class Regla:
    """Una regla de relación: si `coincide` es cierto para un producto,
    `buscar` devuelve una lista de productos candidatos (puede haber
    huecos/None si no se encuentra un candidato claro para ese hueco
    concreto — se filtran después)."""
    def __init__(self, nombre, area, coincide, buscar, max_sugerencias=3):
        self.nombre = nombre
        self.area = area  # área a la que aplica esta regla
        self.coincide = coincide
        self.buscar = buscar
        self.max_sugerencias = max_sugerencias


def construir_indices(productos):
    """Índices auxiliares para búsquedas rápidas: por área+subfamilia,
    y todos los productos activos (sin fecha_baja) por área."""
    por_area = defaultdict(list)
    por_area_subfamilia = defaultdict(list)
    for p in productos:
        if p.get('fecha_baja'):
            continue  # no sugerir productos dados de baja
        area = p.get('area', '')
        por_area[area].append(p)
        sf = sin_acentos_mayus(p.get('subfamilia', ''))
        if sf:
            por_area_subfamilia[(area, sf)].append(p)
    return por_area, por_area_subfamilia


def buscar_por_subfamilia(catalogo_area, *nombres_subfamilia):
    """Primer producto cuya subfamilia coincida (sin acentos/mayúsculas)
    con alguno de los nombres dados."""
    objetivo = {sin_acentos_mayus(n) for n in nombres_subfamilia}
    for p in catalogo_area:
        if sin_acentos_mayus(p.get('subfamilia', '')) in objetivo:
            return p
    return None


def buscar_por_keyword(catalogo_area, *palabras, excluir=()):
    """Primer producto cuyo nombre contenga TODAS las palabras dadas
    (búsqueda simple, sin regex) y NINGUNA de las palabras a excluir."""
    for p in catalogo_area:
        n = sin_acentos_mayus(p.get('nombre', ''))
        if all(w in n for w in palabras) and not any(w in n for w in excluir):
            return p
    return None


def misma_marca(nombre_producto, catalogo_area, *nombres_subfamilia, excluir_ref=None):
    """Busca en las subfamilias dadas un producto de la MISMA marca que
    el producto original (primera palabra del nombre, heurística simple
    pero razonable para este catálogo — la marca casi siempre va
    primero, ej. 'AGRADO CHAMPU...', 'NIVEA GEL...')."""
    marca = sin_acentos_mayus(nombre_producto).split()[0] if nombre_producto else ''
    if not marca:
        return None
    objetivo = {sin_acentos_mayus(n) for n in nombres_subfamilia}
    for p in catalogo_area:
        if excluir_ref and p.get('ref') == excluir_ref:
            continue
        if sin_acentos_mayus(p.get('subfamilia', '')) not in objetivo:
            continue
        if sin_acentos_mayus(p.get('nombre', '')).startswith(marca):
            return p
    return None


# ══════════════════════════════════════════════════════════════════════
# REGLAS — DROGUERÍA
# (11 primeras porteadas de REGLAS_COMPRA_CONJUNTA_DROGUERIA en
# buscador.html, para no perder cobertura ya validada; el resto son
# nuevas, a partir de los ejemplos dados)
# ══════════════════════════════════════════════════════════════════════

def reglas_drogueria(por_area, por_area_subfamilia):
    cat = por_area['drogueria']

    def r_lavavajillas(p):
        n = sin_acentos_mayus(p.get('nombre', ''))
        return 'LAVAVAJILLAS' in n and sin_acentos_mayus(p.get('area','')) != 'INDUSTRIAL'

    def buscar_lavavajillas(p):
        return [
            buscar_por_keyword(cat, 'GUANTE'),
            buscar_por_subfamilia(cat, 'Estropajos', 'Con esponja', 'Fibra metálica', 'Bayetas y paños'),
            # Nuevo, a partir del ejemplo del usuario: papel para secar manos
            buscar_por_keyword(cat, 'PAPEL', 'SECAMANOS') or buscar_por_subfamilia(cat, 'Secadores de manos'),
        ]

    def r_fregasuelos(p):
        n = sin_acentos_mayus(p.get('nombre', ''))
        sf = sin_acentos_mayus(p.get('subfamilia', ''))
        return 'FREGASUELOS' in n or 'FRIEGASUELOS' in n or 'LIMPIADORES SUELOS' in sf

    def buscar_fregasuelos(p):
        # A partir del ejemplo del usuario: cubo Y fregona (antes solo
        # sugería fregona + cepillo/recogedor)
        return [
            buscar_por_subfamilia(cat, 'Cubos y accesorios'),
            buscar_por_keyword(cat, 'FREGONA', excluir=('RECAMBIO',)),
        ]

    def r_limpiacristales(p):
        return 'LIMPIACRISTALES' in sin_acentos_mayus(p.get('subfamilia', ''))

    def buscar_limpiacristales(p):
        return [buscar_por_subfamilia(cat, 'Bayetas microfibra')]

    def r_detergente_ropa(p):
        return 'DETERGENTE' in sin_acentos_mayus(p.get('subfamilia', ''))

    def buscar_detergente_ropa(p):
        return [buscar_por_subfamilia(cat, 'Suavizante normal', 'Suavizante profesional', 'Suavizante concentrado')]

    def r_lejia_desinfectante(p):
        n = sin_acentos_mayus(p.get('nombre', ''))
        sf = sin_acentos_mayus(p.get('subfamilia', ''))
        return 'LEJIA' in n or 'DESINFECTANTE' in sf

    def buscar_lejia_desinfectante(p):
        return [buscar_por_keyword(cat, 'GUANTE')]

    def r_multiusos_desengrasante(p):
        sf = sin_acentos_mayus(p.get('subfamilia', ''))
        return 'MULTIUSOS HOGAR' in sf or 'DESENGRASANTE' in sf

    def buscar_multiusos_desengrasante(p):
        return [buscar_por_subfamilia(cat, 'Bayetas y paños')]

    def r_avispas(p):
        return 'AVISPA' in sin_acentos_mayus(p.get('nombre', ''))

    def buscar_avispas(p):
        return [buscar_por_keyword(cat, 'TRAMPA', 'AVISPA')]

    def r_moscas(p):
        return 'MOSCA' in sin_acentos_mayus(p.get('nombre', ''))

    def buscar_moscas(p):
        return [buscar_por_keyword(cat, 'TRAMPA', 'MOSCA') or buscar_por_keyword(cat, 'CEBO', 'MOSCA')]

    def r_cucarachas(p):
        n = sin_acentos_mayus(p.get('nombre', ''))
        return 'CUCARACHA' in n or 'CUCA' in n

    def buscar_cucarachas(p):
        return [buscar_por_keyword(cat, 'TRAMPA', 'CUCARACHA')]

    def r_hormigas(p):
        return 'HORMIGA' in sin_acentos_mayus(p.get('nombre', ''))

    def buscar_hormigas(p):
        return [buscar_por_keyword(cat, 'TRAMPA', 'HORMIGA')]

    # ── Nuevas, a partir de los ejemplos del usuario ──

    def r_cepillo_barrer(p):
        return 'CEPILLO' in sin_acentos_mayus(p.get('nombre', '')) and 'BARRER' in sin_acentos_mayus(p.get('nombre', ''))

    def buscar_cepillo_barrer(p):
        return [
            buscar_por_keyword(cat, 'RECOGEDOR'),
            buscar_por_keyword(cat, 'BOLSA', 'BASURA'),
        ]

    def r_papelera_cubo_basura(p):
        n = sin_acentos_mayus(p.get('nombre', ''))
        return ('CUBO' in n and 'BASURA' in n) or 'PAPELERA' in n

    def buscar_papelera_cubo_basura(p):
        return [buscar_por_keyword(cat, 'BOLSA', 'BASURA')]

    return [
        Regla('drogueria_lavavajillas', 'drogueria', r_lavavajillas, buscar_lavavajillas),
        Regla('drogueria_fregasuelos', 'drogueria', r_fregasuelos, buscar_fregasuelos),
        Regla('drogueria_limpiacristales', 'drogueria', r_limpiacristales, buscar_limpiacristales),
        Regla('drogueria_detergente_ropa', 'drogueria', r_detergente_ropa, buscar_detergente_ropa),
        Regla('drogueria_lejia_desinfectante', 'drogueria', r_lejia_desinfectante, buscar_lejia_desinfectante),
        Regla('drogueria_multiusos_desengrasante', 'drogueria', r_multiusos_desengrasante, buscar_multiusos_desengrasante),
        Regla('drogueria_avispas', 'drogueria', r_avispas, buscar_avispas),
        Regla('drogueria_moscas', 'drogueria', r_moscas, buscar_moscas),
        Regla('drogueria_cucarachas', 'drogueria', r_cucarachas, buscar_cucarachas),
        Regla('drogueria_hormigas', 'drogueria', r_hormigas, buscar_hormigas),
        Regla('drogueria_cepillo_barrer', 'drogueria', r_cepillo_barrer, buscar_cepillo_barrer),
        Regla('drogueria_cubo_basura', 'drogueria', r_papelera_cubo_basura, buscar_papelera_cubo_basura),
    ]


# ══════════════════════════════════════════════════════════════════════
# REGLAS — PERFUMERÍA
# Todas nuevas, a partir de los ejemplos del usuario y extensiones
# razonables del mismo patrón dentro de la propia área.
# ══════════════════════════════════════════════════════════════════════

def reglas_perfumeria(por_area, por_area_subfamilia):
    cat = por_area['perfumeria']

    def r_champu(p):
        sf = sin_acentos_mayus(p.get('subfamilia', ''))
        return sf.startswith('CHAMPU')

    def buscar_champu(p):
        # Mascarilla de la misma marca si existe, y un cepillo del pelo
        # (tal como pidió el usuario explícitamente) — los cepillos no
        # tienen subfamilia propia en el Sheet (van como "General"), así
        # que se buscan por palabra clave en el nombre.
        return [
            misma_marca(p.get('nombre',''), cat, 'Mascarilla capilar', excluir_ref=p.get('ref'))
                or buscar_por_subfamilia(cat, 'Mascarilla capilar'),
            buscar_por_keyword(cat, 'CEPILLO'),
        ]

    def r_gel_ducha(p):
        return sin_acentos_mayus(p.get('subfamilia', '')) == 'GEL DE DUCHA'

    def buscar_gel_ducha(p):
        return [
            buscar_por_keyword(cat, 'ESPONJA'),
            buscar_por_subfamilia(cat, 'Otros desodorantes', 'Roll-on', 'Stick')
                if not buscar_por_keyword(cat, 'ESPONJA') else
                misma_marca(p.get('nombre',''), cat, 'Otros desodorantes', 'Roll-on', 'Stick', excluir_ref=p.get('ref'))
                or buscar_por_subfamilia(cat, 'Otros desodorantes', 'Roll-on', 'Stick'),
        ]

    def r_espuma_afeitar(p):
        n = sin_acentos_mayus(p.get('nombre', ''))
        return 'AFEITAR' in n or 'AFEITADO' in n

    def buscar_espuma_afeitar(p):
        return [
            buscar_por_keyword(cat, 'MAQUINILLA') or buscar_por_keyword(cat, 'CUCHILLA', 'AFEITAR'),
            buscar_por_subfamilia(cat, 'After shave'),
        ]

    def r_pasta_dental(p):
        return sin_acentos_mayus(p.get('subfamilia', '')) == 'PASTA DENTAL'

    def buscar_pasta_dental(p):
        return [buscar_por_subfamilia(cat, 'Enjuague bucal / Elixir')]

    def r_tinte(p):
        sf = sin_acentos_mayus(p.get('subfamilia', ''))
        return 'TINTE' in sf or sf == 'DECOLORACION'

    def buscar_tinte(p):
        return [
            buscar_por_subfamilia(cat, 'Mascarilla capilar', 'Tratamiento reparador'),
        ]

    def r_crema_facial(p):
        sf = sin_acentos_mayus(p.get('subfamilia', ''))
        return sf in ('CREMA FACIAL DIA', 'CREMA FACIAL NOCHE')

    def buscar_crema_facial(p):
        return [
            buscar_por_subfamilia(cat, 'Contorno de ojos'),
            buscar_por_subfamilia(cat, 'Sérum y tratamientos'),
        ]

    def r_eau_toilette(p):
        sf = sin_acentos_mayus(p.get('subfamilia', ''))
        return sf.startswith('EAU DE')

    def buscar_eau_toilette(p):
        # Desodorante de la misma marca si existe — el emparejamiento
        # perfume+desodorante de la misma gama es una venta cruzada
        # habitual en perfumería.
        return [
            misma_marca(p.get('nombre',''), cat, 'Otros desodorantes', 'Roll-on', 'Stick', excluir_ref=p.get('ref')),
        ]

    def r_gomina_laca(p):
        sf = sin_acentos_mayus(p.get('subfamilia', ''))
        return sf in ('GOMINA Y GEL', 'LACA PARA EL PELO', 'ESPUMA / MOUSSE', 'CERA Y POMADA')

    def buscar_gomina_laca(p):
        return [buscar_por_subfamilia(cat, 'Champú uso frecuente')]

    return [
        Regla('perfumeria_champu', 'perfumeria', r_champu, buscar_champu),
        Regla('perfumeria_gel_ducha', 'perfumeria', r_gel_ducha, buscar_gel_ducha),
        Regla('perfumeria_espuma_afeitar', 'perfumeria', r_espuma_afeitar, buscar_espuma_afeitar),
        Regla('perfumeria_pasta_dental', 'perfumeria', r_pasta_dental, buscar_pasta_dental),
        Regla('perfumeria_tinte', 'perfumeria', r_tinte, buscar_tinte),
        Regla('perfumeria_crema_facial', 'perfumeria', r_crema_facial, buscar_crema_facial),
        Regla('perfumeria_eau_toilette', 'perfumeria', r_eau_toilette, buscar_eau_toilette),
        Regla('perfumeria_gomina_laca', 'perfumeria', r_gomina_laca, buscar_gomina_laca),
    ]


# ══════════════════════════════════════════════════════════════════════
# REGLAS — PINTURAS
# Simplificación en forma de regla de la lógica ya existente en
# productosComplementarios() de buscador.html.
# ══════════════════════════════════════════════════════════════════════

def reglas_pinturas(por_area, por_area_subfamilia):
    cat = por_area['pinturas']

    def es_base_disolvente(p):
        n = sin_acentos_mayus(p.get('nombre', ''))
        return bool(re.search(r'\bESMALTE\b|\bOXIRON\b|\bSINTETIC|\bBARNIZ\b|\bLACA\b', n))

    def r_pintura_base(p):
        n = sin_acentos_mayus(p.get('nombre', ''))
        return es_base_disolvente(p) or 'PLASTIC' in n

    def buscar_pintura_base(p):
        sugerencias = []
        if es_base_disolvente(p):
            sugerencias.append(buscar_por_keyword(cat, 'DISOLVENTE', 'UNIVERSAL'))
        sugerencias.append(buscar_por_keyword(cat, 'CINTA', 'ENMASCARAR'))
        sugerencias.append(buscar_por_keyword(cat, 'PAPEL', 'PROTECTOR'))
        return sugerencias

    return [
        Regla('pinturas_base', 'pinturas', r_pintura_base, buscar_pintura_base),
    ]


def generar_sugerencias(productos):
    por_area, por_area_subfamilia = construir_indices(productos)
    todas_las_reglas = reglas_drogueria(por_area, por_area_subfamilia) \
        + reglas_perfumeria(por_area, por_area_subfamilia) \
        + reglas_pinturas(por_area, por_area_subfamilia)

    reglas_por_area = defaultdict(list)
    for r in todas_las_reglas:
        reglas_por_area[r.area].append(r)

    filas = []
    contador_reglas = defaultdict(int)

    for p in productos:
        if p.get('fecha_baja'):
            continue
        if p.get('relacionados'):
            continue  # ya tiene informado algo — no se sugiere encima, se respeta lo existente
        area = p.get('area', '')
        for regla in reglas_por_area.get(area, []):
            if not regla.coincide(p):
                continue
            candidatos = [c for c in regla.buscar(p) if c and c.get('ref') != p.get('ref')]
            # quitar duplicados conservando el orden
            vistos, unicos = set(), []
            for c in candidatos:
                if c['ref'] not in vistos:
                    vistos.add(c['ref'])
                    unicos.append(c)
            unicos = unicos[:regla.max_sugerencias]
            if not unicos:
                continue
            contador_reglas[regla.nombre] += 1
            filas.append({
                'referencia': p.get('ref', ''),
                'nombre': p.get('nombre', ''),
                'area': area,
                'familia': p.get('familia', ''),
                'subfamilia': p.get('subfamilia', ''),
                'regla': regla.nombre,
                'relacionados_sugeridos': ','.join(c['ref'] for c in unicos),
                'relacionados_nombres': ' | '.join(c['nombre'] for c in unicos),
            })
            break  # una sola regla por producto, la primera que coincida

    return filas, contador_reglas


def exportar_excel(filas, ruta_salida):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Sugerencias'

    cabeceras = ['referencia', 'nombre', 'area', 'familia', 'subfamilia',
                 'regla', 'relacionados_sugeridos', 'relacionados_nombres', 'aprobado (SI/NO)']
    ws.append(cabeceras)
    for cell in ws[1]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill('solid', fgColor='1A1A1A')
        cell.alignment = Alignment(horizontal='center')

    for fila in filas:
        ws.append([
            fila['referencia'], fila['nombre'], fila['area'], fila['familia'], fila['subfamilia'],
            fila['regla'], fila['relacionados_sugeridos'], fila['relacionados_nombres'], '',
        ])

    anchos = [16, 42, 12, 22, 26, 26, 26, 46, 14]
    for i, ancho in enumerate(anchos, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = ancho

    ws.freeze_panes = 'A2'
    wb.save(ruta_salida)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--entrada', default='productos.json', help='Ruta al productos.json de entrada')
    ap.add_argument('--salida', default='sugerencias_relacionados.xlsx', help='Ruta del Excel de salida')
    args = ap.parse_args()

    with open(args.entrada, encoding='utf-8') as f:
        datos = json.load(f)
    productos = datos.get('productos', datos)

    print(f'Cargados {len(productos)} productos de {args.entrada}')

    filas, contador_reglas = generar_sugerencias(productos)

    print(f'\nSugerencias generadas: {len(filas)}')
    print('\nPor regla:')
    for nombre_regla, n in sorted(contador_reglas.items(), key=lambda x: -x[1]):
        print(f'  {nombre_regla}: {n}')

    exportar_excel(filas, args.salida)
    print(f'\n✓ Excel de revisión generado: {args.salida}')
    print('  Revisa la columna "relacionados_nombres", marca "aprobado" donde corresponda,')
    print('  y pega "referencia" + "relacionados_sugeridos" (para las filas aprobadas) en la')
    print('  columna "relacionados" de la hoja Productos del Sheet.')


if __name__ == '__main__':
    main()
